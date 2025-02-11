import openpyxl

# Caminho do arquivo
caminho_arquivo = ""

# Lista de abas ocultas que queremos verificar
abas_ocultas = {}

# Abrir o arquivo Excel
wb = openpyxl.load_workbook(caminho_arquivo, data_only=False, read_only=True)  # Manter as fórmulas

# Obter todas as planilhas visíveis
abas_visiveis = [sheet for sheet in wb.sheetnames if wb[sheet].sheet_state == "visible"]

# Criar um conjunto para armazenar abas ocultas que são referenciadas
abas_referenciadas = set()

processo = '['
# Percorrer todas as abas visíveis e verificar fórmulas
for aba_nome in abas_visiveis:
    processo += '#'
    print(f'Processo de leitura: {processo}')
    aba = wb[aba_nome]
    for row in aba.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):  # Apenas fórmulas
                for aba_oculta in abas_ocultas:
                    if f"{aba_oculta.upper()}!" in cell.value.upper():  # Verifica referência à aba oculta
                        abas_referenciadas.add(aba_oculta)


processo += ']'
print(f'Processo de leitura: {processo}')

# Encontrar as abas ocultas que não foram referenciadas
abas_inuteis = abas_ocultas - abas_referenciadas

# Exibir os resultados
print("Abas ocultas referenciadas:", abas_referenciadas)
print("Abas ocultas que parecem inúteis:", abas_inuteis)
